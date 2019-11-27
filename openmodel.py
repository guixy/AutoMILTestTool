from PyQt5.QtWidgets import  *
import sys

import shutil

from openui import Ui_MainWindow

from PyQt5.QtCore import *
from PyQt5.QtGui import *
import  os
import  re

import time
import TestOperation
import matlab.engine
import createcase
import xlrd

from openpyxl import load_workbook
from threading import Thread

class EmittingStr(QObject):
    textWritten = pyqtSignal(str) #定义一个发送str的信号
    def write(self, text):
      self.textWritten.emit(str(text))

class Oprate(QThread):
    sig = pyqtSignal()
    def __init__(self, parent=None):
        super(Oprate, self).__init__(parent)
    def run(self):
        self.sig.emit()






class basePage(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super(basePage, self).__init__()
        self.setupUi(self)
        self.Select_btn.clicked.connect(self.SelectModelAndInitMatlab)

        self.Init_btn.toggle()

        self.Init_btn.clicked.connect(lambda :self.CheckBtn(1))
        self.Static_btn.clicked.connect(lambda :self.CheckBtn(2))
        self.InitCase_btn.clicked.connect(lambda :self.CheckBtn(3))
        self.GerCase_btn.clicked.connect(lambda :self.CheckBtn(4))
        self.Start_btn.clicked.connect(lambda :self.CheckBtn(5))
        self.barlabel = QLabel()
        self.statusBar.addPermanentWidget(self.barlabel)
        sys.stdout = EmittingStr(textWritten=self.outputWritten)
        self.Oprate=Oprate()
        self.Oprate.sig.connect(self.InitModel)
        self.creatcases=createcase.Ui_MainWindow()
        self.eng = matlab.engine.start_matlab()
        self.barlabel.setText('未选择模型' )



    def ChooseProDir(self):
        dir=QFileDialog.getExistingDirectory()
        dir=dir.replace('/','\\')
        #self.ProjectName_2.setText(dir)
        self.LibrP=dir

        if dir=='':

            sys.exit(0)



        #self.eng = matlab.engine.start_matlab('-desktop')
            #self.adds(dir,self.child0)
        #a.initUI()
    def SelectModelAndInitMatlab(self):
        dir1, file1 = QFileDialog.getOpenFileName(filter="*.slx")

        (filepath, tempfilename) = os.path.split(dir1)
        (filename, extension) = os.path.splitext(tempfilename)
        dir = filepath.replace('/', '\\')
        #self.Libr=os.path.
        self.modelname=filename[4:]
        if filename!="":
            self.barlabel.setText('选择的模型是：'+filename)
            self.TO = TestOperation.Mlab(filename, dir, self.eng,self.LibrP)


    def CheckBtn(self,num):
        self.Select_btn.setEnabled(False)
        self.Init_btn.setEnabled(False)
        self.Static_btn.setEnabled(False)
        self.InitCase_btn.setEnabled(False)
        self.GerCase_btn.setEnabled(False)
        self.Start_btn.setEnabled(False)
        if num==1:
            self.Num=1
            t = Thread(target=self.InitModel)
            t.start()
        elif num==2:
            t = Thread(target=self.StaticCheck)
            t.start()
            self.Num=2
        elif num==3:
            t = Thread(target=self.InitCase)
            t.start()
            self.Num=3
        elif num==4:
            #t = Thread(target=self.GerCases)
            #t.start()
            self.GerCases()
            self.Num=4
        elif num==5:
            t = Thread(target=self.StartTestCase)
            t.start()


    def CheckNum(self):
        if self.Num==1:
            self.InitModel()
            self.Num = 0
        elif self.Num==2:
            self.StaticCheck()
            self.Num = 0
        elif self.Num==3:
            self.InitCase()
            self.Num = 0
        elif self.Num==4:
            self.GerCases()
            self.Num=0

    def InitModel(self):
        try:
            if self.barlabel.text()=="" :
                print("未选择模型！")
            else:
                print("初始化开始！")
                self.TO.platform()

                self.TO.InitM()
                print("初始化成功！")

            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)


        except Exception as e:
            print("初始化失败！请从err.log查看具体错误信息！")
            print(str(e))

            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)

    def StaticCheck(self):
        try:

            self.TO.StaticCheck()
            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)
            print("静态测试成功！")
        except Exception as e:
            print("静态测试失败！请从err.log查看具体错误信息！")
            print(str(e))
            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)

    def InitCase(self):
        try:
            self.TO.InitCase()
            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)
            print("初始化模板成功！")
        except Exception as e:
            print("初始化模板失败！请从err.log查看具体错误信息！")
            print(str(e))
            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)

    def GerCases(self):
        try:

            self.di = QMainWindow()

            self.creatUI = self.creatcases

            self.creatUI.setupUi(self.di)

            self.creatUI.pushButton.clicked.connect(self.WriteXlSX)
            self.creatUI.pushButton_2.clicked.connect(self.Save)



            a,b=self.readXLSX()


            if a=='a':
                print('找不到测试用例模板')
                self.Select_btn.setEnabled(True)
                self.Init_btn.setEnabled(True)
                self.Static_btn.setEnabled(True)
                self.InitCase_btn.setEnabled(True)
                self.GerCase_btn.setEnabled(True)
                self.Start_btn.setEnabled(True)
            #b=['1','2','3']
            else:
                self.creatUI.comboBox.clear()

                self.creatUI.comboBox_2.clear()
                self.creatUI.comboBox.addItems(a)
                self.creatUI.comboBox_2.addItems(b)

                self.di.show()
                self.Select_btn.setEnabled(True)
                self.Init_btn.setEnabled(True)
                self.Static_btn.setEnabled(True)
                self.InitCase_btn.setEnabled(True)
                self.GerCase_btn.setEnabled(True)
                self.Start_btn.setEnabled(True)
        except Exception as e:
                print(str(e))
                self.Select_btn.setEnabled(True)
                self.Init_btn.setEnabled(True)
                self.Static_btn.setEnabled(True)
                self.InitCase_btn.setEnabled(True)
                self.GerCase_btn.setEnabled(True)
                self.Start_btn.setEnabled(True)
    def StartTestCase(self):
        try:
            self.eng.cd(self.LibrP)
            self.TO.StartTest()
            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)
            print("Harness测试成功！")
        except Exception as e:
            print("Harness测试失败！请从err.log查看具体错误信息！")
            print(str(e))
            self.Select_btn.setEnabled(True)
            self.Init_btn.setEnabled(True)
            self.Static_btn.setEnabled(True)
            self.InitCase_btn.setEnabled(True)
            self.GerCase_btn.setEnabled(True)
            self.Start_btn.setEnabled(True)

    def readXLSX(self):

        #list = os.listdir(os.path.join(os.getcwd(), 'testLibrary/UnitTest/'))
        list = os.listdir( self.LibrP+'/UnitTest/')

        #workbook1111 = xlrd.open_workbook(os.path.join(os.getcwd(),'1.'))
        self.PATH=""
        for i in list:

            if self.modelname[4:] in i:
                #list2 = os.listdir(os.path.join(os.getcwd(), 'testLibrary/UnitTest/' + i))
                list2 = os.listdir( self.LibrP+'/UnitTest/' + i)
                for j in list2:
                    if '001_Spec.xlsx' in j and "~$"not in j:
                        #path=os.path.join(os.getcwd(), 'testLibrary/UnitTest/' + i)
                        path = self.LibrP+'/UnitTest/' + i
                        self.PATH=path+'/'+j
                        #print(self.PATH)

                        workbook1=xlrd.open_workbook(filename=self.PATH)
        self.tempfile = self.PATH[:len(self.PATH) - 5] + 'a.xlsx'



        if self.PATH!="":
            sheet = workbook1.sheets()[0]

            #sheet = workbook1.sheet_by_name(name)
            row = sheet.row_values(1)
            row1=sheet.row_values(0)

            n=0
            signales=[]
            self.singalesDIC={}
            for k1 in row1:
                if k1=='Input':
                    n=n+1
            self.N=10+n
            for k in range(11,11+n-1):

                signales.append(row[k])
                self.singalesDIC[row[k]]=k

            col=sheet.col_values(10)
            self.colDIC={}
            times=[]

            for k3 in range(1,len(col)):
                if isinstance(col[k3],float):

                    times.append(str(col[k3]))
                    self.colDIC[col[k3]]=k3
            #self.time=times
            col2=sheet.col_values(0)

            needSig=[]
            for k4 in range(2,len(col2)):
                if col2[k4]!="":

                    needSig.append(col2[k4])
            if needSig :

                self.sigs=needSig
                self.AddNeedSig(needSig)
                self.AddExp(needSig,times)



            workbook1.release_resources()


            return signales,times
        else:
            return "a",'b'

    def WriteXlSX(self):

        i=self.creatUI.comboBox.currentIndex()
        txt=self.creatUI.comboBox.itemText(i)
        j = self.creatUI.comboBox_2.currentIndex()
        txt1 = self.creatUI.comboBox_2.itemText(j)
        inputVal=self.creatUI.inputVal.text()
        #exVal=self.creatUI.expVal.text()
        if inputVal=="" :
            QMessageBox.about(self, "消息", "输入值为空！")
        else:
            '''workbook = xlrd.open_workbook(self.PATH)
            workbooknew = copy(workbook)
            ws = workbooknew.get_sheet(0)'''
            len(self.singalesDIC)
            #tempp = re.split('.xlsx', self.PATH)[0]
            temp = self.PATH
            '''if os.path.exists(temp):
                pass
            else:
                shutil.copyfile(self.PATH,temp)'''
            #temp='D:\工作\工作\脚本\matlab+python\dist\XMiMatlGerTest/aaa.xlsx'
            if os.path.exists(self.tempfile):
                wb = load_workbook(self.tempfile)
            else:
                wb = load_workbook(self.PATH)
            wb1 = wb.active

            #print(self.colDIC[float(txt1)], self.singalesDIC[txt])
            wb1.cell(self.colDIC[float(txt1)]+1, self.singalesDIC[txt]+1, float(inputVal))
            #wb1.cell(self.colDIC[float(txt1)]+1, self.N+1, float(exVal))
            '''wb.write(self.colDIC[float(txt1)], self.singalesDIC[txt], self.creatUI.inputVal.text())
            wb.write(self.colDIC[float(txt1)], self.N, self.creatUI.expVal.text())'''

            wb.save(self.tempfile)

            #shutil.copyfile(temp,temp[:len(temp)-4]+'zip')
            #shutil.copyfile(temp[:len(temp )-4] + 'zip',temp )
            #os.rename(temp,temp[:len(temp)-4]+'_templat.zip')
            #os.remove(temp)
            #os.rename(temp[:len(temp)-5]+'a.xlsx',temp)
            #print('填入'+txt+"在"+txt1+"时刻的值为："+inputVal+"   期望值为："+exVal)
            print('填入' + txt + "在" + txt1 + "时刻的值为：" + inputVal )

    def AddNeedSig1(self,sigs):
        labelnames=locals()
        horizonLay=locals()
        linetext = locals()
        linetext1=locals()
        linetext2 = locals()
        linetext3 = locals()
        linetext4 = locals()
        linetext5 = locals()
        layout=QGridLayout(self.creatUI.scrollArea)
        label1=QLabel("误差")
        label2 = QLabel("变化率上限")
        label3 = QLabel("变化率下限")
        label4 = QLabel("最大值")
        label5 = QLabel("最小值")
        layout.addWidget(label1, 0, 1)
        layout.addWidget(label2, 0, 2)
        layout.addWidget(label3, 0, 3)
        layout.addWidget(label4, 0, 4)
        layout.addWidget(label5, 0, 5)
        layout.setSpacing(10)
        sigs=["5","","4","3","2","1"]
        for i in range(0,len(sigs)):
            #exec('self.btn{} = {}'.format(i, QLabel()))
            '''layout['self.layoutWidget'+str(i)] = QWidget(self.creatUI.scrollArea)
            layout['self.layoutWidget' + str(i)].setGeometry(QRect(10, 140, 751, 31))
            layout['self.layoutWidget' + str(i)].setObjectName("layoutWidget"+str(i))
            horizonLay['self.creatUI.horizontalLayout' + str(i)] = QHBoxLayout(self.creatUI.scrollArea)
            horizonLay['self.creatUI.horizontalLayout' + str(i)].setContentsMargins(0, 0, 0, 0)
            horizonLay['self.creatUI.horizontalLayout' + str(i)].setObjectName("horizontalLayout" + str(i))'''

            labelnames['self.label'+str(i)]=QLabel(sigs[i])
            linetext['self.btn1'+str(i)]=QLineEdit()
            linetext['self.btn2' + str(i)] = QLineEdit()
            linetext['self.btn3' + str(i)] = QLineEdit()
            linetext['self.btn4' + str(i)] = QLineEdit()
            linetext['self.btn5' + str(i)] = QLineEdit()

            layout.addWidget(labelnames['self.label'+str(i)], i + 1, 0)
            layout.addWidget(linetext['self.btn1' + str(i)], i + 1, 1)
            layout.addWidget(linetext['self.btn2' + str(i)], i + 1, 2)
            layout.addWidget(linetext['self.btn3' + str(i)], i + 1, 3)
            layout.addWidget(linetext['self.btn4' + str(i)], i + 1, 4)
            layout.addWidget(linetext['self.btn5' + str(i)], i + 1, 5)


            #horizonLay['self.creatUI.horizontalLayout' + str(i)].addWidget(labelnames['self.label' + str(i)])
            '''horizonLay['self.creatUI.horizontalLayout'+str(i)] = QHBoxLayout(layout['self.creatUI.layoutWidget' + str(i)])
            horizonLay['self.creatUI.horizontalLayout' + str(i)].setContentsMargins(0, 0, 0, 0)
            horizonLay['self.creatUI.horizontalLayout' + str(i)].setObjectName("horizontalLayout"+str(i))
            self.label11 = QLabel(layout['self.creatUI.layoutWidget' + str(i)])
            self.label11.setObjectName("label")
            self.label11.setText('1212312312')'''
            #horizonLay['self.creatUI.horizontalLayout' + str(i)].addWidget(self.label)



            '''labelnames['self.creatUI.label' +str(i)] = QLabel(layout['self.creatUI.layoutWidget' + str(i)])
            labelnames['self.creatUI.label' +str(i)].setObjectName("label"+str(i))
            labelnames['self.creatUI.label' + str(i)].setText("信号：")
            horizonLay['self.creatUI.horizontalLayout' + str(i)].addWidget(labelnames['self.creatUI.label'+str(i)])'''


            #self.sig_label = QLabel()

            #self.creatUI.scrollArea.addPermanentWidget(self.btn_i)

    def AddNeedSig(self,sigs):
        #sigs = ["5", "22", "4", "3", "2", "1"]
        self.table1=QTableWidget(len(sigs),5)
        # TODO 优化 2 设置水平方向表格为自适应的伸缩模式
        self.table1.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table1.setHorizontalHeaderLabels(['误差', '变化率上限', '变化率下限','最大值','最小值'])
        #table.setFixedWidth(700)
        #for i in sigs:

        self.table1.setVerticalHeaderLabels(sigs)
        '''for i in range(0,len(sigs)):
            newItem1 = QLineEdit()
            table.setCellWidget(i, 0, newItem1)
            newItem2 = QLineEdit()
            table.setCellWidget(i, 1, newItem2)
            newItem3 = QLineEdit()
            table.setCellWidget(i, 2, newItem3)
            newItem4 = QLineEdit()
            table.setCellWidget(i, 3, newItem4)
            newItem5 = QLineEdit()
            table.setCellWidget(i, 4, newItem5)'''
        self.creatUI.scrollArea.setWidget(self.table1)
    def AddExp(self,sigs,times):
        #sigs=['1','2','2']
        self.table2 = QTableWidget(len(sigs), 4)
        self.table2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table2.setHorizontalHeaderLabels(['期望信号', '时刻', '期望值',""])

        cb = QComboBox()

        cb.addItems(times)
        self.table2.setSpan(0, 1, len(sigs), 1)
        self.table2.setCellWidget(0, 1, cb)

        for i in range(0,len(sigs)):

            newItem = QTableWidgetItem(sigs[i])
            self.table2.setItem(i,0,newItem)
            #num = QTableWidgetItem('0')
            #table.setItem(i,2,num)

            self.table2.setSpan(0,3,len(sigs),1)
        self.btnexp=QPushButton()
        self.btnexp.setText("确定")
        self.table2.setCellWidget(0,3,self.btnexp)
        self.btnexp.clicked.connect(self.WritrEXP)
        self.creatUI.scrollArea_2.setWidget(self.table2)

    def WritrEXP(self):
        if os.path.exists(self.tempfile):
            wb = load_workbook(self.tempfile)
        else:
            wb = load_workbook(self.PATH)
        wb1 = wb.active
        for ii in range(0,len(self.sigs)):
            exp = self.table2.item(ii,2)
            txt1=self.table2.cellWidget(0, 1).currentIndex()
            txt1=self.table2.cellWidget(0,1).itemText(txt1)
            if exp:
                expVal=exp.text()
                wb1.cell(self.colDIC[float(txt1)] + 1, self.N + 1+ii, float(expVal))
                print('填入信号'+self.sigs[ii]  + "在" + txt1 + "时刻的值为：" + expVal )
        wb.save(self.tempfile)
            #else:
                #QMessageBox.about(self, "消息", "为空！")
            #print(self.table2.cellWidget(ii,1).currentIndex())
    def Save(self):
        if os.path.exists(self.tempfile):
            wb = load_workbook(self.tempfile)
            wb1 = wb.active
            for ii in range(0, len(self.sigs)):
                for jj in range(0,5):
                    valnum = self.table1.item(ii,jj)
                    if valnum:
                        val=valnum.text()
                        wb1.cell(3 + ii,  2+ jj, float(val))
                    else:
                        QMessageBox.about(self, "消息", "请输入必填部分数值！")
                        break
            wb.save(self.tempfile)
            os.remove(self.PATH)
            os.rename(self.tempfile,self.PATH)


        else:
           pass
        self.di.close()

    def outputWritten(self, text):
        cursor = self.textBrowser.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(text)
        self.textBrowser.setTextCursor(cursor)
        self.textBrowser.ensureCursorVisible()
